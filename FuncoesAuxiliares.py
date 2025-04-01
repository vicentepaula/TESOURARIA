import pygetwindow as gw
import cv2
import pyautogui
import time
import datetime
import csv
from decimal import Decimal
import locale
import tkinter as tk
from tkinter import messagebox
import logging
from pywinauto import Application
import Calculos
import FuncoesAuxiliares
import pandas as pd
import RepositorioDAO
from openpyxl import load_workbook
import os



class Funcao_Apoio:

  
    #Retorna o título da Janela Ativa
    def GetScreenShot(self,caminhoArquivo, nmLoja):
        screenshot = pyautogui.screenshot()
        data_hora_atual = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        nome_arquivo = f'{nmLoja}_{data_hora_atual}.png'
        
        screenshot.save(f"{caminhoArquivo}{nome_arquivo}")


    def capturar_primeira_data(self, loja):

        arquivo =  f"\\\\10.11.10.3\\arcomixfs$\\Financeiro\\digitacao\\movimentolj{loja}.xlsx" 

        # Abre o arquivo em modo somente leitura para melhor desempenho
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active  # Pega a aba ativa automaticamente

        # Percorre a coluna de índice 3 (quarta coluna no Excel)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4, values_only=True):
            if row[0]:  # Se encontrar uma célula preenchida
                wb.close()  # Fecha o arquivo para liberar memória
                return row[0].strftime("%d/%m/%Y")  # Retorna a data formatada

        wb.close()
        return None  # Retorna None se não encontrar nenhuma data
    
   
    

    def check_window_exists(self, window_name):
        time.sleep(2)
        # Obtém todas as janelas ativas
        windows = gw.getWindowsWithTitle(window_name)
        
        # Verifica se a lista de janelas não está vazia
        if windows:
            return True
        else:
            return False

    

    # Aguarda até a janela abrir.
    def AguardaAberturaJanela(self, janela_alvo):

     # Nome da janela que estou está esperando abrir
       
           # Loop para verificar continuamente se a janela alvo foi aberta
        while True:
            # Obtém todas as janelas ativas
            janelas = gw.getAllWindows()

            # Verifica se a janela alvo está entre as janelas ativas
            if any(janela_alvo in janela.title for janela in janelas):
                print("Janela alvo foi aberta!:")
                break  # Sai do loop quando a janela alvo for encontrada

            # Pausa por um curto período de tempo antes de verificar novamente
            time.sleep(1)  # Importe time se você ai

       
               
   
     # Testa se a janela existe baseado em imagens
    def aguardar_janela_por_imagem(self,imagem_janela, mensagem): # Passar o endereço da imagem salvo no computador
        
        while True:

            try:
                posicao = pyautogui.locateCenterOnScreen(imagem_janela, confidence=0.2)
                if posicao is not None:
                   return 0
          
            except Exception as e:
                print(f'Aguardando {mensagem}...', e)

    

    def show_popup(self,message):
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal
        messagebox.showinfo("Popup", message)
        root.destroy()

    def GeraLogsInfo(self, mensagem):
        logging.basicConfig(level=logging.INFO, filename="C:\\Projetos_Python\\TESOURARIA\\arquivos\\logs\\Info\\info_processamento.log")
        logging.info(mensagem)

    def InsertOutros(self, valor, controle_box):
        guptamdiframeAcer_Detalhamento = Application().connect(title_re=".*Movimento Detalhado.*")
        guptadialog = guptamdiframeAcer_Detalhamento[u'Gupta:Dialog']
        guptadialog.Wait('ready')
        guptachildtable = guptadialog[u'Gupta:ChildTable']
      #  guptachildtable.click_input() # Clica na tabela

        varFuncao = FuncoesAuxiliares.Funcao_Apoio()
        varcalculos = Calculos.Operacoes()
        valor = varcalculos.removeCifraoRetornaString(valor)
    
        pyautogui.sleep(0.35)
        pyautogui.press('ctrl')
        pyautogui.sleep(0.35)
        pyautogui.press('insert') # Abre a linha para digitação 
        pyautogui.sleep(0.35)
        pyautogui.press('insert') # Inseri a linha 
      #  if controle_box >= "1": #if controle_box == "1" or controle_box == "2":
     #        pyautogui.press('insert') # Abre a linha 
        pyautogui.press('insert') # Abre a linha 
        pyautogui.sleep(0.35)
        pyautogui.write("787 - ")
        pyautogui.press('down')
        pyautogui.sleep(0.35)
        pyautogui.press('up')
        pyautogui.sleep(0.35)
        pyautogui.press('tab')
        pyautogui.sleep(0.35)
        pyautogui.write(valor) # Informando o  valor 
        
        guptadialog[u'Button3'].click_input() #Click na no  botão conciliar
        pyautogui.sleep(2)
        
        jnErro = varFuncao.check_window_exists("Atenção") # Se abrir uma caixa de diálogo com o nome atenção
        if jnErro == True:
            appDlgConf_Exclusao = Application().connect(title_re=".*Atenção.*")
            window = appDlgConf_Exclusao.Dialog
            window.Wait('ready')
            button = window[u'&OK']
            button.click_input()
            
            varFuncao.esperar_fechamento_janela("Atenção")
          
            appDlgConf_Exclusao = Application().connect(title_re=".*Movimento Detalhado.*")
            window = appDlgConf_Exclusao.Dialog
            window.Wait('ready')
            button = window[u'&OK']
            window.close()

            varFuncao.esperar_fechamento_janela("Movimento Detalhado")

            jnErro = varFuncao.check_window_exists("Atenção") # Se abrir uma caixa de diálogo com o nome atenção
            if jnErro == True:
                appDlgConf_Exclusao = Application().connect(title_re=".*Atenção.*")
                window = appDlgConf_Exclusao.Dialog
                window.Wait('ready')
                button = window[u'&OK']
                button.click_input()
        else:
           guptadialog[u'Button6'].double_click_input() # Confirma e fecha a janela


    def InsertGeral(self):
        guptamdiframeAcer_Detalhamento = Application().connect(title_re=".*Movimento Detalhado.*")
        guptadialog = guptamdiframeAcer_Detalhamento[u'Gupta:Dialog']
        guptadialog.Wait('ready')
                           
        guptadialog[u'Button3'].click_input() #Click na no  botão conciliar              [u'Button3', u'Concilia Todos', u'Concilia TodosButton']
        pyautogui.sleep(2)

        guptadialog[u'Button6'].click_input() # Confirma e fecha a janela
    
        
    
    """
    def Trt_lj987(self):
        
        time.sleep(3)
        pyautogui.click(pyautogui.locateCenterOnScreen("C:\\Projetos_Python\\FISCAL\\Img\\campos_janela_auxiliar\\cbox_cabo_janela_auxililar.png",confidence=0.9), duration=0.15)
        time.sleep(1)
        pyautogui.scroll(-300)
        time.sleep(1)
        pyautogui.rightClick(pyautogui.locateCenterOnScreen("C:\\Projetos_Python\\FISCAL\\Img\\campos_janela_auxiliar\\cbox_central987_janela_auxiliar.png",confidence=0.85), duration=0.15)
        time.sleep(1)
        pyautogui.click(pyautogui.locateCenterOnScreen("C:\\Projetos_Python\\FISCAL\\Img\\campos_janela_auxiliar\\btn_desmarcarTodos.png",confidence=0.85), duration=0.15)
        time.sleep(1)
        pyautogui.click(pyautogui.locateCenterOnScreen("C:\\Projetos_Python\\FISCAL\\Img\\campos_janela_auxiliar\\cbox_central987_janela_auxiliar.png",confidence=0.85), duration=0.15)
        time.sleep(1)
    """

    
    def esperar_fechamento_janela(self, janela_alvo):
    
        
        # Loop para verificar continuamente se a janela alvo foi aberta
        while True:
            # Obtém todas as janelas ativas
            janelas = gw.getAllWindows()

            # Verifica se a janela alvo está entre as janelas ativas
            if any(janela_alvo in janela.title for janela in janelas):
                print("Janela alvo foi aberta!")
            else:
                print("Janela alvo não existe mais. Encerrando o script.")
                break  # Sai do loop quando a janela alvo não existe mais

            # Pausa por um curto período de tempo antes de verificar novamente
            time.sleep(1)  # Importe time se você ainda não o fez

    """
    def lerCsv(self):
        # Abre o arquivo CSV em modo de leitura
        with open('C:\\Projetos_Python\\TESOURARIA\\arquivos\\digitacao\\movimentolj27.csv') as f:
         next(f)

         for line in f:
              line=line.strip()
              line=line.split(",")
              print(" Dados da Planilha : ",line)
   

    def lerCsv_01(self):
        # Abre o arquivo CSV em modo de leitura
        with open('C:\\Projetos_Python\\TESOURARIA\\arquivos\\digitacao\\movimentolj27.csv', newline='') as csvfile:
        # Cria um leitor CSV
         reader = csv.reader(csvfile)

        # Itera sobre as linhas do arquivo
        for row in reader:
            print(', '.join(row))  # Imprime cada linha como uma lista separada por vírgulas

    """


    def SeJanelaExiste_porImagem(self, imagem_janela):
        time.sleep(3)
        try:
                posicao = pyautogui.locateCenterOnScreen(imagem_janela, confidence=0.2)
                if posicao is not None:
                   return True
          
        except Exception as e:
                return False

    

    def converter_para_decimal(self, valor_com_cifrao):
        valor_sem_cifrao=None
        #Removendo o ponto da casa de milhar se houver
        valor_com_cifrao = valor_com_cifrao.replace('.','').lstrip()

        # Verificar se o valor possui cifrão
        if 'R$' in valor_com_cifrao:
            # Remover o cifrão
            valor_sem_cifrao = valor_com_cifrao.replace('R$','').lstrip()
                           
            try:
                #Removendo a virgula da casa decimal e inserindo o ponto para que possa fazer a conversão
                valor_sem_cifrao = valor_sem_cifrao.replace(',','.')

                # Converter para decimal
                valor_decimal = Decimal(valor_sem_cifrao)

                return valor_decimal
            except ValueError:
                print("Erro: Valor não pôde ser convertido para decimal.")
                return None
        else:
            try: 
                valor_sem_cifrao = valor_com_cifrao
                #Removendo a virgula da casa decimal e inserindo o ponto para que possa fazer a conversão
                valor_sem_cifrao = valor_sem_cifrao.replace(',','.')                             
                #fazendo a conversao                
                valor_decimal = Decimal(valor_sem_cifrao)
                
                return valor_decimal
            
            except ValueError:
                print(f"Erro: Valor não pôde ser convertido para decimal AGORAs.",ValueError)
                return None

        
    def SelecionaConteudoCampo(self):
        pyautogui.keyDown('ctrl')
        pyautogui.keyDown('a')
        pyautogui.keyUp('a')
        pyautogui.keyUp('ctrl')


    def copiarCampo(self):
        pyautogui.keyDown('ctrl')
        pyautogui.keyDown('c')
        pyautogui.keyUp('c')
        pyautogui.keyUp('ctrl')
    
        
    def subtracao(self, vlr1, vlr2):
        if vlr1 =="":
            vlr1==0
        if vlr2 =="":
            vlr2==0
                    
        return vlr1 - vlr2
        

    def xlsx_to_csv(self, nro_lj):
                     
        #xlsx_file= f'C:/Projetos_Python/TESOURARIA/arquivos/download/movimentolj{nro_lj}.xlsx'

        xlsx_file = f"\\\\10.11.10.3\\arcomixfs$\\Financeiro\\digitacao\\movimentolj{nro_lj}.xlsx"
        csv_file =  f'C:\\Projetos_Python\\TESOURARIA\\arquivos\\execucao\\movimentolj{nro_lj}.csv'
       
        
        workbook = load_workbook(filename=xlsx_file, data_only=True)
        # Obter o nome da planilha ativa
        active_sheet_name = workbook.active.title

        # Ler o arquivo .xlsx
        df = pd.read_excel(xlsx_file, engine='openpyxl', sheet_name= active_sheet_name, header=0)

            
        # Iterar sobre todas as colunas para garantir que os números e datas sejam tratados corretamente
        for i, col in enumerate(df.columns):
            # Tratamento específico para a quarta coluna (índice 3)

            if i == 0 or i == 8 or i == 9 or i == 10:
                #print(df[col].astype)
                # Verificar se o valor pode ser convertido em float, ignorando strings inválidas
                df[col] = df[col].apply(
            lambda x: str(int(float(x)))  # Converte para int e depois para string, removendo as casas decimais
            if pd.notna(x) and isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).isdigit()
            else str(x)  # Retorna o valor original convertido para string se não for numérico
            )
                  
            if i == 2:
          #      print(df[col].astype)
                # Verificar se o valor pode ser convertido em float, ignorando strings inválidas
                df[col] = df[col].apply(lambda x:f'{int(float(x)):,}'.replace(',', 'X').replace('.', ',').replace('X', '.')  # Converte float para int e formata
                if pd.notna(x) and isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).isdigit()  # Verifica se o valor é numérico
                else x  # Retorna o valor original se não puder ser convertido
                )

            elif i == 3:
               # print("")
                df[col] = df[col].apply(lambda x: 
                pd.to_datetime(x).strftime('%d/%m/%y')  # Converte a data para o formato DD-MM-YY
                if pd.notna(x) and pd.to_datetime(x, errors='coerce') is not pd.NaT  # Verifica se é uma data válida
                else str(x)  # Se não for uma data, converte para string
                )
            
            elif i == 4:
                df[col] = df[col].apply(
                lambda x: f'{float(x):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                if pd.notna(x) and isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).isdigit()
                else x  # Retorna o valor original se não for numérico
                )
            
            elif df.shape[1] > 11:
               df.drop(df.columns[11], axis=1, inplace=True)
                    
        df.to_csv(csv_file, sep=';', index=False, encoding='ISO-8859-1', decimal=',', na_rep='', header=True)    


    def verifica_e_apaga_arquivo(self, caminho_arquivo):
        # Verifica se o arquivo existe
        if os.path.exists(caminho_arquivo):
            try:
                # Apaga o arquivo
                os.remove(caminho_arquivo)
                print(f"Arquivo '{caminho_arquivo}' apagado com sucesso.")
            except Exception as e:
                print(f"Erro ao tentar apagar o arquivo: {e}")
        else:
            print(f"O arquivo '{caminho_arquivo}' não existe.")

    #___________________________________________________________________________________________________________________________________________
    

    def GeraSeqTurnoCSV(self, nroLoja, dtaMovimento):
        # Carregando o CSV e garantindo que todas as colunas sejam tratadas como strings
        df_csv = pd.read_csv(
            f'C:\\Projetos_Python\\TESOURARIA\\arquivos\\execucao\\movimentolj{nroLoja}.csv',  
            sep=';', 
            encoding='ISO-8859-1',
            dtype=str  # Força todas as colunas a serem lidas como strings
        )

        varExecute = RepositorioDAO.DAO()
        con = varExecute.getConection()

        print(df_csv.columns)  
        if "SEQTURNO" not in df_csv.columns:
            sql_query = f""" 
                SELECT h.seqturno, TO_CHAR(h.dtamovimento, 'DD/MM/YYYY') as Data, h.coo, 
                    h.nroempresa || h.nrocheckout || h.coo as KEY
                FROM consincomonitor.tb_docto h
                WHERE h.nroempresa = '{nroLoja}' 
                AND h.dtamovimento = TO_DATE('{dtaMovimento}', 'DD/MM/YYYY')
                """  
            
            df_sql = pd.read_sql_query(sql_query, con.conectar())

            #Garantindo que a coluna 'KEY' seja tratada como string em ambos os DataFrames
            df_csv['KEY'] = df_csv['KEY'].astype(str)
            df_sql['KEY'] = df_sql['KEY'].astype(str)
                        
            
            # Realizando o merge
            df_merged = pd.merge(df_csv, df_sql, on='KEY', how='left')
            print(df_merged.columns)

            #Convertendo a coluna 'seqturno' para string, se existir no DataFrame
            df_merged = df_merged.convert_dtypes(str)
            
            # Salvando o DataFrame resultante no CSV, com todos os dados como string
            df_merged.to_csv(
                f'C:\\Projetos_Python\\TESOURARIA\\arquivos\\execucao\\movimentolj{nroLoja}.csv', 
                index=False, 
                sep=";", 
                encoding='ISO-8859-1', 
                na_rep='',  # Para evitar 'NaN' nos campos vazios
            )

   
    def excluir_coluna_11(self, nro_lj):
        caminho_entrada = f"\\\\10.11.10.3\\arcomixfs$\\Financeiro\\digitacao\\movimentolj{nro_lj}.xlsx"
        caminho_saida = f"\\\\10.11.10.3\\arcomixfs$\\Financeiro\\digitacao\\movimentolj{nro_lj}.xlsx"

        workbook = load_workbook(filename=caminho_entrada, data_only=True)
        # Obter o nome da planilha ativa
        active_sheet_name = workbook.active.title

        # Ler o arquivo .xlsx
        df = pd.read_excel(caminho_entrada, engine='openpyxl', sheet_name= active_sheet_name, header=0)


        # Carregar a planilha
        #df = pd.read_excel(caminho_entrada)

        # Verificar se a coluna 11 existe
        if df.shape[1] > 11:
            df.drop(df.columns[11], axis=1, inplace=True)

        # Salvar a planilha sem a coluna
        df.to_excel(caminho_saida, index=False)

       
    #Retorna o título da Janela Ativa
    def GetTituloJanelaAtiva(self):
        janela_ativa = gw.getActiveWindow().title
        return janela_ativa
           

   
    
            
                           


            
                
                 